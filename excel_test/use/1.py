import pandas as pd
from openpyxl import load_workbook
import shutil
import os

# 定义数据表和模板文件路径
data_file = '2024年收文记录.xls'
template_file = '2024年收文簿.xls'

# 创建输出目录
output_dir = 'OA收文簿'
os.makedirs(output_dir, exist_ok=True)

# 指定要操作的数据表和模板中的工作表名称
data_sheet_name = 'OA收文'  # 数据表中要使用的工作表名称
template_sheet_name = 'OA'  # 模板表中要使用的工作表名称

# 读取数据表格中的指定工作表
data = pd.read_excel(data_file, sheet_name=data_sheet_name)


# 获取用户输入的行号列表
input_indices = input("请输入要使用的行号（用逗号分隔，如 0,1,2）：")
row_indices = [int(x.strip()) for x in input_indices.split(',')]

# 遍历指定的行号
for row_index in row_indices:
    if row_index < 0 or row_index >= len(data):
        print(f"警告: 行号 {row_index} 超出范围，跳过此行。")
        continue

    # 获取指定行的数据
    row = data.iloc[row_index]

    # 复制模板文件
    output_file = os.path.join(output_dir, f'Template_{row["Name"]}.xlsx')
    shutil.copy(template_file, output_file)

    # 加载复制的模板文件中的指定工作表
    workbook = load_workbook(output_file)

    if template_sheet_name not in workbook.sheetnames:
        print(f"警告: 模板工作表 {template_sheet_name} 不存在，跳过此文件。")
        continue

    sheet = workbook[template_sheet_name]

    # 准备句子模板
    sentence_template = "收文2024年第{name}号"

    # 生成完整的句子
    sentence = sentence_template.format(name=row['编号'])

    # 填充数据到模板
    sheet['A2'] = row['编号']
    sheet['A4'] = row['来文单位']
    sheet['D4'] = row['收文文号']
    sheet['A6'] = row['内容摘要']

    # 保存填充数据后的文件
    workbook.save(output_file)

    print(f"模板文件已使用第 {row_index + 1} 行的数据生成：{output_file}")

print("所有指定行的数据模板文件生成完成。")
